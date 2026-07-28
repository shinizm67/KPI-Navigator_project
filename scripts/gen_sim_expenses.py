#!/usr/bin/env python3
"""検証用シミュレーション支出データ生成。

各年の日次収入(xlsx)を読み、支出を「収入に対する比率」で生成する。
これにより収入の単位に依存せず FL・利益・コスト構造が整合する。

出力（excel/ 配下、雛形と同一フォーマット）:
  - 検証用_{year}支出_日次.csv  : 日付=YYYY-MM-DD,費目,金額（食材/ドリンク/アルバイト）
  - 検証用_{year}支出_月次.csv  : 日付=YYYY-MM,費目,金額（固定+その他11費目）

支出はすべて「収入のある営業日/月」にのみ計上する。
"""
import csv
import os
import random
from collections import defaultdict

import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
EXCEL = os.path.join(os.path.dirname(HERE), "excel")

INCOME_FILES = {
    2024: "2024年売上入力用.xlsx",
    2025: "2025年売上入力用.xlsx",
    2026: "2026年売上入力用のコピー.xlsx",
}

# --- 日次変動費（対 当日売上） ---
DAILY_RATIOS = [
    ("食材仕入れ費", 0.25),
    ("ドリンク仕入れ費", 0.08),
    ("アルバイト人件費", 0.18),
]
DAILY_NOISE = 0.06  # ±6% の日々ゆらぎ

# --- 月次：固定費（対 年平均月次売上 = 月ごと一定） ---
FIXED_RATIOS = [
    ("家賃", 0.08),
    ("固定人件費", 0.12),
    ("損害保険", 0.004),
    ("通信費", 0.003),
]

# --- 月次：変動・準変動費（対 当月売上、光熱費は季節係数） ---
def elec_season(m):
    if m in (7, 8, 9):
        return 1.25
    if m in (12, 1, 2):
        return 1.15
    return 1.0


def gas_season(m):
    if m in (12, 1, 2, 3):
        return 1.30
    if m in (7, 8):
        return 0.80
    return 1.0


VAR_MONTHLY = [
    ("備品・消耗品仕入費", lambda ms, m: 0.012 * ms),
    ("雑費・小口精算費", lambda ms, m: 0.008 * ms),
    ("電気代", lambda ms, m: 0.025 * ms * elec_season(m)),
    ("ガス代", lambda ms, m: 0.018 * ms * gas_season(m)),
    ("水道代", lambda ms, m: 0.008 * ms),
    ("広告宣伝費", lambda ms, m: 0.015 * ms),
    ("クレジットカード手数料", lambda ms, m: 0.020 * ms),
]


def read_income(path):
    """returns list of (date, sales_float) for days with sales present."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]
    rows = []
    for r in range(2, ws.max_row + 1):
        d = ws.cell(row=r, column=1).value
        sales = ws.cell(row=r, column=4).value
        if d is None or sales is None:
            continue
        rows.append((d.date(), float(sales)))
    return rows


def gen_year(year, income_path):
    rng = random.Random(year * 1000 + 7)
    days = read_income(income_path)
    # 月次売上・年平均月次売上
    month_sales = defaultdict(float)
    month_present = set()
    for d, s in days:
        month_sales[d.month] += s
        month_present.add(d.month)
    ams = (sum(month_sales.values()) / len(month_present)) if month_present else 0.0

    # ---- 日次 ----
    daily_rows = []
    for d, s in days:
        iso = d.strftime("%Y-%m-%d")
        for label, ratio in DAILY_RATIOS:
            noise = 1.0 + rng.uniform(-DAILY_NOISE, DAILY_NOISE)
            amt = int(round(s * ratio * noise))
            if amt < 1:
                amt = 1
            daily_rows.append((iso, label, amt))

    # ---- 月次 ----
    monthly_rows = []
    for m in sorted(month_present):
        ym = "%d-%02d" % (year, m)
        ms = month_sales[m]
        for label, ratio in FIXED_RATIOS:
            amt = int(round(ams * ratio))
            monthly_rows.append((ym, label, max(amt, 1)))
        for label, fn in VAR_MONTHLY:
            amt = int(round(fn(ms, m)))
            monthly_rows.append((ym, label, max(amt, 1)))

    return daily_rows, monthly_rows, month_sales, ams


def write_csv(path, rows):
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(["日付", "費目", "金額"])
        w.writerows(rows)


def main():
    print("year | ann.sales | ann.exp | profit | margin% | FL%(var) ")
    for year, fname in INCOME_FILES.items():
        path = os.path.join(EXCEL, fname)
        daily_rows, monthly_rows, month_sales, ams = gen_year(year, path)
        d_out = os.path.join(EXCEL, "検証用_%d支出_日次.csv" % year)
        m_out = os.path.join(EXCEL, "検証用_%d支出_月次.csv" % year)
        write_csv(d_out, daily_rows)
        write_csv(m_out, monthly_rows)

        ann_sales = sum(month_sales.values())
        daily_exp = sum(a for _, _, a in daily_rows)
        monthly_exp = sum(a for _, _, a in monthly_rows)
        ann_exp = daily_exp + monthly_exp
        profit = ann_sales - ann_exp
        margin = (profit / ann_sales * 100) if ann_sales else 0
        flpct = (daily_exp / ann_sales * 100) if ann_sales else 0
        print(
            "%d | %9.0f | %9.0f | %8.0f | %6.1f | %6.1f"
            % (year, ann_sales, ann_exp, profit, margin, flpct)
        )
        print("     -> %s (%d rows)" % (os.path.basename(d_out), len(daily_rows)))
        print("     -> %s (%d rows)" % (os.path.basename(m_out), len(monthly_rows)))


if __name__ == "__main__":
    main()
